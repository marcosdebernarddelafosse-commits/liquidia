"""Exportación del papel de trabajo de IVA a Excel.

Genera el documento que el contador archiva y usa para cargar la DDJJ en AFIP:
encabezado del período, apertura de ventas y compras por alícuota, y el cuadro
de liquidación con el saldo resultante.

Usa openpyxl directamente para controlar el formato (un papel de trabajo prolijo
vende el producto tanto como el cálculo correcto).
"""

from __future__ import annotations

from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .liquidacion import LiquidacionIVA

# Paleta sobria.
_AZUL = "1F3864"
_GRIS = "D9D9D9"
_VERDE = "C6EFCE"
_ROJO = "FFC7CE"

_TITULO = Font(bold=True, color="FFFFFF", size=12)
_NEGRITA = Font(bold=True)
_MONEDA = '#,##0.00'
_BORDE = Border(*(Side(style="thin", color="BFBFBF"),) * 4)


def _encabezado(ws, fila: int, texto: str, ncols: int = 4) -> int:
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=ncols)
    celda = ws.cell(row=fila, column=1, value=texto)
    celda.font = _TITULO
    celda.fill = PatternFill("solid", fgColor=_AZUL)
    celda.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[fila].height = 20
    return fila + 1


def _fila_titulos(ws, fila: int, titulos: list[str]) -> int:
    for col, t in enumerate(titulos, start=1):
        c = ws.cell(row=fila, column=col, value=t)
        c.font = _NEGRITA
        c.fill = PatternFill("solid", fgColor=_GRIS)
        c.border = _BORDE
        c.alignment = Alignment(horizontal="center")
    return fila + 1


def _renglones_alicuota(ws, fila: int, por_alicuota: dict) -> int:
    for ali in sorted(por_alicuota, reverse=True):
        r = por_alicuota[ali]
        ws.cell(row=fila, column=1, value=f"{ali}%").border = _BORDE
        for col, val in ((2, r.neto), (3, r.iva)):
            c = ws.cell(row=fila, column=col, value=float(val))
            c.number_format = _MONEDA
            c.border = _BORDE
        fila += 1
    return fila


def _fila_total(ws, fila: int, etiqueta: str, valor: Decimal,
                resaltar: str | None = None) -> int:
    c1 = ws.cell(row=fila, column=1, value=etiqueta)
    c1.font = _NEGRITA
    c2 = ws.cell(row=fila, column=3, value=float(valor))
    c2.font = _NEGRITA
    c2.number_format = _MONEDA
    if resaltar:
        for col in (1, 2, 3):
            ws.cell(row=fila, column=col).fill = PatternFill("solid", fgColor=resaltar)
    return fila + 1


def exportar_papel_de_trabajo(liq: LiquidacionIVA, ruta: str) -> str:
    """Escribe el papel de trabajo de IVA en `ruta` (xlsx). Devuelve la ruta."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Liquidación IVA"

    fila = 1
    fila = _encabezado(ws, fila, f"LIQUIDACIÓN DE IVA — Período {liq.periodo}")
    ws.cell(row=fila, column=1, value=f"CUIT: {liq.cuit}").font = _NEGRITA
    fila += 2

    # Débito fiscal (ventas)
    fila = _encabezado(ws, fila, "DÉBITO FISCAL — Ventas")
    fila = _fila_titulos(ws, fila, ["Alícuota", "Neto Gravado", "IVA"])
    fila = _renglones_alicuota(ws, fila, liq.ventas_por_alicuota)
    fila = _fila_total(ws, fila, "Total Débito Fiscal", liq.debito_fiscal, _GRIS)
    fila += 1

    # Crédito fiscal (compras)
    fila = _encabezado(ws, fila, "CRÉDITO FISCAL — Compras")
    fila = _fila_titulos(ws, fila, ["Alícuota", "Neto Gravado", "IVA"])
    fila = _renglones_alicuota(ws, fila, liq.compras_por_alicuota)
    fila = _fila_total(ws, fila, "Total Crédito Fiscal", liq.credito_fiscal, _GRIS)
    fila += 1

    # Cuadro de liquidación
    fila = _encabezado(ws, fila, "DETERMINACIÓN DEL SALDO")
    fila = _fila_total(ws, fila, "Débito Fiscal", liq.debito_fiscal)
    fila = _fila_total(ws, fila, "(-) Crédito Fiscal", liq.credito_fiscal)
    fila = _fila_total(ws, fila, "= Saldo Técnico", liq.saldo_tecnico)
    fila = _fila_total(ws, fila, "(-) Percepciones IVA", liq.percepciones_iva)
    if liq.saldo_a_pagar > 0:
        fila = _fila_total(ws, fila, "SALDO A PAGAR", liq.saldo_a_pagar, _ROJO)
    else:
        fila = _fila_total(ws, fila, "SALDO A FAVOR", liq.saldo_a_favor, _VERDE)

    # Advertencias de calidad
    if liq.advertencias:
        fila += 1
        fila = _encabezado(ws, fila, "ADVERTENCIAS DE CONTROL")
        for adv in liq.advertencias:
            ws.cell(row=fila, column=1, value=f"• {adv}")
            ws.merge_cells(start_row=fila, start_column=1,
                           end_row=fila, end_column=4)
            fila += 1

    # Anchos de columna
    for col, ancho in ((1, 28), (2, 18), (3, 18), (4, 12)):
        ws.column_dimensions[get_column_letter(col)].width = ancho

    Path(ruta).parent.mkdir(parents=True, exist_ok=True)
    wb.save(ruta)
    return ruta
